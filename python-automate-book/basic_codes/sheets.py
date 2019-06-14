***************************************************************
****************  Lendo documentos excel  *********************
***************************************************************

---------------------------------------------------------------------------------------------------------------
|wb = openpyxl.load_workbook('example.xlsx') 	|	a Funcao openpyxl.load_workbook() recebe o nome do arquivo e retorna um valor
|												|	com o tipo de dado workbook. Esse objeto Workbook representa o arquivo Excel,
|												|	de modo um pouco semelhante a maneira como um objeto File representa um arquivo texto aberto.
-----------------------------------------------------------------------------------------------------------
|wb.get_sheet_names()							|	retorna as folhas planilhas abertas dentro do documento	
-------------------------------------------------------------------------------------------
|wb.get_sheet_by_name()							|	Cada planilha eh representada por um objeto Worksheet, que pode ser obtido se passarmos
|wb.get_active_sheet()							|	a string com o nome da planilha para o metodo get_sheet_by_name() do workbook. Por fim,
|												|	podemos chamar o metodo get_active_sheet() de um objeto Workbook para obter a planilha ativa do workbook.
|												|	A planilha ativa eh aquela que estara em evidencia quando o workbook for aberto no Excel. De posse do objeto 
|												|	Worksheet, voce podera obter o seu nome a partir do atributo title.								|
-----------------------------------------------------------------------------------

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
sheet.max_row
sheet.max_column


Podemos determinar o tamanho da planilha usando os metodos .max_column e .max_row do objeto Worksheet.




*******************************************************************************
********************* Abrindo e obtendo informacoes tecnicas ******************
*******************************************************************************


	Cada planilha eh representada por um objeto Worksheet, que pode ser obtido se passarmos
	a string com o nome da planilha para o metodo get_sheet_by_name() do workbook. Por fim,
	podemos chamar o metodo get_active_sheet() de um objeto Workbook para obter a planilha ativa do workbook.
	A planilha ativa eh aquela que estara em evidencia quando o workbook for aberto no Excel. De posse do objeto 
	Worksheet, voce podera obter o seu nome a partir do atributo title.


import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet3')
print(sheet)
print(type(sheet))
print(sheet.title)
anotherSheet = wb.get_active_sheet()
print(anotherSheet)


------------------------------------------------------------------------------------------------------------

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
sheet.max_row
sheet.max_column


Podemos saber o tamanho da planilha usando os metodos .max_column e .max_row do objeto Worksheet.

------------------------------------------------------------------------------------------

Especificar a coluna pela letra eh complicado para programar porque apos a coluna Z
as colunas comecam a usar letras: AA, AB, AC e assim por diante


Para fazer a conversao de letras para numeros, chame a funcao column_index_frin_string()
Para converter de numeros para letras chame a funcao get_column_letter()

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))

print(get_column_letter(1))
print(get_column_letter(27))
print(get_column_letter(900))


*******************************************************************************
****************  obtendo informacoes e valores no documento ******************
*******************************************************************************

	O objeto Cell tem um atributo value que contem o valor armazenado nessa celula. Os
	objetos Cell tambem tem atributos row, column e coordinate que fornecem informacoes sobre a localizacao da celula.
	Nesse caso, acessar o atributo value de nosso objeto Cell para a celula B1 nos da a string 'Apples'. O atributo row
	nos fornece o inteiro 1, o atributo column nos fornece 'B' e o atributo coordinate nos da 'B1'

	O openpyxl interpretara automaticamente as datas na coluna A e as retornara como valores datetime em vez de strings


import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet['A1'])
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)

print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
print('Cell ' + c.coordinate + ' is ' + c.value)
print(sheet['C1'].value)


print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1,8,2):
	print(i, sheet.cell(row=1, column=2).value)


	Como alternativa, podemos tambem obter uma celula usando o metodo cell() da planilha e passando inteiros
	para seus argumentos nomeados row e column. O inteiro correspondente a primeira linha ou coluna eh 1, e nao 0. 

	Como podemos ver, usar o metodo cell() da planilha e passar row=1 e column=2 fara voce obter um objeto Cell para a celula B1,
	assim como ocorreu quando especificamos sheet['B1'].
	Entao, ao usar o metodo cell() e seus argumentos nomeados, podemos criar um loop for para exibir os valores de uma serie de celulas.
	
	A variavel i do loop for eh passada para o argumento nomeado row do metodo cell(), enquanto 2 eh sempre passado para o argumento
	nomeado column. 

*******************************************************************************
****************  obtendo linhas e colunas das planilhas **********************
*******************************************************************************

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
#print(tuple(sheet['A1':'C3']))

for rowOfCellObjects in sheet['A1':'C3']:
	for cellObj in rowOfCellObjects:
		print(cellObj.coordinate, cellObj.value)
	print('--- END OF ROW ---')

'''
Nesse caso, definimos que queremos objetos Cell da area retangular de A1 a C3 e obtivemos
um objeto Generator contendo os objetos Cell dessa area. Para nos ajudar a visualizar esse objeto 
Generator, podemos usar tuple() nesse objeto de modo a exibir seus objetos Cell em uma tupla.
Essa tupla contem tres tuplas, uma para cada linha, da parte superior da area desejada ate a parte inferior.
Cada uma dessas tres tuplas internas contem objetos Cell de uma linha da area desejada, da celula mais a esquerda para a celula mais a direita
De modo geral, nosso slice da planilha contem todos os objetos Cell da area de A1 a C3, comecando na celula superior a esquerda terminando na celula inferior a direita
Para exibir os valores de cada celula dessa area, utilizamos dois loops for, O loop for externo percorre todas as linhas do slice. Entao cada linha, o loop for aninhado percorre todas as celulas dessa linha
'''
-----------------------------------------------------------------------------
-----------------------------------------------------------------------------

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_active_sheet()
for cellObj in sheet["B"]:
	print(cellObj.value)

'''
sheet.columns[1] substituido por sheet["B"]
Ao usar o atributo rows de um objeto Worksheet, vamos obter uma tupla de tuplas.
Cada uma dessas tuplas internas representa uma linha e contem os objetos Cell dessa linha.
O atributo columns tambem fornece uma tupla de tuplas, em que cada uma das tuplas internas contem objetos Cell de uma coluna em particular.
Em example.xlsx, como ha 7 linhas e 3 colunas, rows fornece uma tupla de 7 tuplas (cada qual contendo 3 objetos Cell) e columns fornece uma tupla de 3 tuplas (cada qual contendo 7 objetos Cell).
'''


-----------------------------------------------------------------------------
-------------------------	PROGRAMAS	-------------------------------------
-----------------------------------------------------------------------------

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet3')
print(sheet)
print(type(sheet))
print(sheet.title)
anotherSheet = wb.get_active_sheet()
print(anotherSheet)

'''
	Cada planilha eh representada por um objeto Worksheet, que pode ser obtido se passarmos
	a string com o nome da planilha para o metodo get_sheet_by_name() do workbook. Por fim,
	podemos chamar o metodo get_active_sheet() de um objeto Workbook para obter a planilha ativa do workbook.
	A planilha ativa eh aquela que estara em evidencia quando o workbook for aberto no Excel. De posse do objeto 
	Worksheet, voce podera obter o seu nome a partir do atributo title.
'''

-----------------------------------------------------------------------------
-----------------------------------------------------------------------------

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
#print(sheet['A1'])
#print(sheet['A1'].value)
c = sheet['B1']
#print(c.value)

print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
print('Cell ' + c.coordinate + ' is ' + c.value)
print(sheet['C1'].value)

'''
	O objeto Cell tem um atributo value que contem o valor armazenado nessa celula. Os
	objetos Cell tambem tem atributos row, column e coordinate que fornecem informacoes sobre a localizacao da celula.
	Nesse caso, acessar o atributo value de nosso objeto Cell para a celula B1 nos da a string 'Apples'. O atributo row
	nos fornece o inteiro 1, o atributo column nos fornece 'B' e o atributo coordinate nos da 'B1'

	O openpyxl interpretara automaticamente as datas na coluna A e as retornara como valores datetime em vez de strings
'''

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1,8,2):
	print(i, sheet.cell(row=1, column=2).value)

'''
	Como alternativa, podemos tambem obter uma celula usando o metodo cell() da planilha e passando inteiros
	para seus argumentos nomeados row e column. O inteiro correspondente a primeira linha ou coluna eh 1, e nao 0. 
'''

-----------------------------------------------------------------------------
-----------------------------------------------------------------------------

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
sheet.max_row
sheet.max_column

'''
	Podemos determinar o tamanho da planilha usando os metodos .max_column e .max_row do objeto Worksheet.
'''

-----------------------------------------------------------------------------
-----------------------------------------------------------------------------


import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))

'''
	Para fazer a conversao de letras para numeros, chame a funcao column_index_frin_string()
	Para converter de numeros para letras chame a funcao get_column_letter()
'''

print(get_column_letter(1))
print(get_column_letter(27))
print(get_column_letter(900))

'''
	Especificar a coluna pela letra eh complicado para programar porque apos a coluna Z
	as colunas comecam a usar letras: AA, AB, AC e assim por diante
'''

-----------------------------------------------------------------------------
-----------------------------------------------------------------------------

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
#print(tuple(sheet['A1':'C3']))

for rowOfCellObjects in sheet['A1':'C3']:
	for cellObj in rowOfCellObjects:
		print(cellObj.coordinate, cellObj.value)
	print('--- END OF ROW ---')

'''
Nesse caso, definimos que queremos objetos Cell da area retangular de A1 a C3 e obtivemos
um objeto Generator contendo os objetos Cell dessa area. Para nos ajudar a visualizar esse objeto 
Generator, podemos usar tuple() nesse objeto de modo a exibir seus objetos Cell em uma tupla.
Essa tupla contem tres tuplas, uma para cada linha, da parte superior da area desejada ate a parte inferior.
Cada uma dessas tres tuplas internas contem objetos Cell de uma linha da area desejada, da celula mais a esquerda para a celula mais a direita
De modo geral, nosso slice da planilha contem todos os objetos Cell da area de A1 a C3, comecando na celula superior a esquerda terminando na celula inferior a direita
Para exibir os valores de cada celula dessa area, utilizamos dois loops for, O loop for externo percorre todas as linhas do slice. Entao cada linha, o loop for aninhado percorre todas as celulas dessa linha
'''

-----------------------------------------------------------------------------
-----------------------------------------------------------------------------

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_active_sheet()
for cellObj in sheet["B"]:
	print(cellObj.value)

'''
sheet.columns[1] substituido por sheet["B"]
Ao usar o atributo rows de um objeto Worksheet, vamos obter uma tupla de tuplas.
Cada uma dessas tuplas internas representa uma linha e contem os objetos Cell dessa linha.
O atributo columns tambem fornece uma tupla de tuplas, em que cada uma das tuplas internas contem objetos Cell de uma coluna em particular.
Em example.xlsx, como ha 7 linhas e 3 colunas, rows fornece uma tupla de 7 tuplas (cada qual contendo 3 objetos Cell) e columns fornece uma tupla de 3 tuplas (cada qual contendo 7 objetos Cell).
'''

-----------------------------------------------------------------------------
-----------------------------------------------------------------------------